import openpyxl

class HomePageData:

    test_HomePage_data = [{"Name":"Carlos","Email": "atank88@hotmail.com","Password":"BadPassword&5050"},
                            {"Name":"Jose", "Email":"decirt48@hotmail.com", "Password":"Password&222"}]

    @staticmethod
    def getTestMethod(testcase):

        book = openpyxl.load_workbook("C:\\Users\\carlos.picon\\Development\\PythonSelfFramework\\TestData\\pythonExcel.xlsx")

        sheet = book.active

        Dict = {}

        for i in range(sheet.max_row):
            if sheet.cell(row=i+1, column=1).value == testcase:
                for n in range(sheet.max_column - 1):
                    Dict[sheet.cell(row=1, column=n + 2).value] = sheet.cell(row=i + 1, column=n + 2).value


        return [Dict]