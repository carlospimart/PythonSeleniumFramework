import openpyxl

book = openpyxl.load_workbook("C:\\Users\\carlos.picon\\Development\\PythonSelfFramework\\TestData\\pythonExcel.xlsx")

sheet = book.active

#how to read
cell = sheet.cell(row=1, column=2)
print(cell.value)

#how to write
#sheet.cell(row=2, column=2).value= "Rahul"

print(sheet.cell(row=2, column=2).value)

#how to get the number of rows
print("maximum row: ",sheet.max_row)

#how to get the number of columns
print(sheet.max_column)

#how to get the value of a specific box
print(sheet["A3"].value)

print("-------------------")

list_dict=[]
#how to get all the values of the sheet
for i in range(sheet.max_row-1):
        Dict = {}
        for n in range(sheet.max_column-1):
            Dict[sheet.cell(row=1, column=n+2).value]=sheet.cell(row=i+2, column=n+2).value
        list_dict.append(Dict)


print(list_dict[0],list_dict[1],list_dict[2])


def test_Method(testcase):

    Dict = {}

    for i in range(sheet.max_row):
        if sheet.cell(row=i + 1, column=1).value == testcase:
            for n in range(sheet.max_column-1):

                Dict[sheet.cell(row=1, column=n + 2).value] = sheet.cell(row=i + 1, column=n + 2).value

    return [Dict]

print(test_Method("Testcase4")[0])